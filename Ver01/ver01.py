from openpyxl import Workbook, load_workbook

# Excel文件路径指定
wkBook = load_workbook("C:\\Users\\Desktop\\oldwu\\text.xlsx")
wkSheet = wkBook["Sheet1"]

# 存放要读取的数据
dic={}
# 股票代码第一行
rowIdx = 4
# 股票代码所在列
idColIdx = 1

# 先读取要设置的股票,暂时存一下
while wkSheet.cell(row=rowIdx, column=idColIdx).value:
  tmpId = wkSheet.cell(row=rowIdx, column=idColIdx).value
  if tmpId.startswith("6"):
    dic[tmpId]=""
  rowIdx = rowIdx + 1

# 读取的txt文件指定
f = open("C:\\Users\\Desktop\\oldwu\\test\\mkdtd00(上海0701X市).txt", "r")

for line in f:
  datas = line.split("|")
  if len(datas) > 4 and datas[1] in dic:
      item ={"code":datas[1],"value1":datas[2],"value2":datas[3]}
      dic[datas[1]]=item

#print(dic)

# 股票代码第一行
rowIdx = 4
# 股票代码所在列
idColIdx = 1
# 输出：成交数量所在列
cntColIdx = 3
# 输出：成交金额所在列
sumColIdx = 4

# 把读取到的数据写到指定列
while wkSheet.cell(row=rowIdx, column=idColIdx).value:
  if wkSheet.cell(row=rowIdx, column=idColIdx).value in dic:
    tItem = dic[wkSheet.cell(row=rowIdx, column=idColIdx).value]
    wkSheet.cell(row=rowIdx, column=cntColIdx).value = tItem["value1"]
    wkSheet.cell(row=rowIdx, column=sumColIdx).value = tItem["value2"]
  rowIdx = rowIdx + 1

# 保存一下
wkBook.save("output.xlsx")

print("End")
